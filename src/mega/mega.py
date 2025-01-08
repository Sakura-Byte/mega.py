from openpyxl import load_workbook
import os
from mega import Mega
from megaapi import get_file_name_from_mega_link

mega = Mega()
mega_client = mega.login(email="pengyoupy001@gmail.com", password="85Ya.mDcTmMGz7y")

wb = load_workbook('DL.xlsx')
ws = wb['Sheet1']

for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    print("Processing row", row[0].value)
    # Assign a new column for the result
    ws.cell(row=row[0].row, column=ws.max_column + 1).value = 'Checked?'
    
    # Initialize flags to track folder creation
    normal_folder_created = False
    mp3_folder_created = False

    mega_link = row[2].value
    if mega_link == '没有MEGA档':
        print("No MEGA link, skipping")
        ws.cell(row=row[0].row, column=ws.max_column + 1).value = '1'
    else:
        mega_links = mega_link.split('|')
        mp3_links = []
        normal_links = []
        for mega_link in mega_links:
            try:
                file_name = get_file_name_from_mega_link(mega_link)
                if "【MP3】" in file_name:
                    print("MP3 link found:", mega_link)
                    mp3_links.append(mega_link)
                else:
                    print("Normal link found:", mega_link)
                    normal_links.append(mega_link)
            except Exception as e:
                print("Error:", e)
        
        # 处理普通链接
        if normal_links:
            path = f'/{row[13].value.split(" - ")[0]}/{row[13].value}/{row[14].value}/{row[0].value}'
            os_prefix = "Z:/115/DL-Back"
            mega_prefix = "/DL-Back"
            # 检查路径是否存在
            if not normal_folder_created and not os.path.exists(os_prefix + path):
                print("Creating normal folder:", mega_prefix + path)
                mega_client.create_folder(mega_prefix + path)
                normal_folder_created = True
            for link in normal_links:
                print("Importing normal link:", link, "to", mega_prefix + path)
                mega_client.import_public_url(link, mega_prefix + path)
        
        # 处理 MP3 链接
        if mp3_links and row[15].value != 'null':
            mp3_folder_num = row[15].value.split(" - ")[0]
            path_mp3 = f'/{row[15].value}/{row[16].value}/{row[0].value}'
            if int(mp3_folder_num) == 1:
                path_mp3 = "/MP3" + path_mp3
            elif int(mp3_folder_num) == 2:
                path_mp3 = "/MP3_02" + path_mp3
            else:
                raise ValueError(f"Invalid mp3 folder number: {mp3_folder_num}")
            
            if not mp3_folder_created and not os.path.exists(os_prefix + path_mp3):
                print("Creating MP3 folder:", mega_prefix + path_mp3)
                mega_client.create_folder(mega_prefix + path_mp3)
                mp3_folder_created = True
            for link in mp3_links:
                print("Importing MP3 link:", link, "to", mega_prefix + path_mp3)
                mega_client.import_public_url(link, mega_prefix + path_mp3)
        
        ws.cell(row=row[0].row, column=ws.max_column + 1).value = '1'

wb.save('DL_result.xlsx')

